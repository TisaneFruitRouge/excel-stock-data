import sys
import openpyxl
import json

def create_transaction_json(shares_dict):
	
	profit_dict = dict()
	transaction_dict = dict()

	for share in shares_dict.keys():
		
		profit_dict[share] = {"profit": 0, "shares_traded": 0}

		share_dict  = shares_dict[share]
		bought_list = share_dict["Bought"]
		sold_list   = share_dict["Sold"]

		profit = 0

		for sell in sold_list: 
			profit_dict[share]["shares_traded"] += sell["shares"]
			profit += sell["shares"]*sell["price"]
			while sell["shares"] > 0:
				for buy in bought_list:
					if sell["shares"]>buy["shares"]:
						# more shares sold than bought
						sell["shares"] -= buy["shares"]
						profit -= buy["shares"]*buy["price"] # adjusting the profit

						if (sell["shares"] < 1e6): # correcting float number problems
							sell["shares"] = 0
						
						buy["shares"] = 0
					else: 
						# more bought sold than sold
						profit -= (buy["shares"] - sell["shares"])*buy["price"] # adjusting the profit
						buy["shares"]-=sell["shares"]
						
						if (buy["shares"] < 1e6): # correcting float number problems
							buy["shares"] = 0
						
						sell["shares"] = 0
						continue
		
		profit_dict[share]["profit"] = profit

	# creating the final dictionnary

	for share in shares_dict.keys():

		transaction_dict[share] = {}
		if not len(shares_dict[share]["Bought"]) == 0:
			transaction_dict[share]["Bought"] = []
		if not len(shares_dict[share]["Sold"]) == 0:	
			transaction_dict[share]["Sold"] = []


		for buy in shares_dict[share]["Bought"]:
			if buy["shares"] == 0:
				continue
			else :
				transaction_dict[share]["Bought"].append(buy)

		for sell in shares_dict[share]["Sold"]:
			if sell["shares"] == 0:
				continue
			else :
				transaction_dict[share]["Sold"].append(sell)

		transaction_dict[share]["profit"] = profit_dict[share]["profit"]

	with open("transactions.json", "w") as json_file:
		json.dump(transaction_dict, json_file, indent=4)

	return profit_dict # return to use in the update_balance_json function


def update_balance_json(profit_dict):
	with open("balance.json", "w+") as json_file:
		try: 
			balance = json.load(json_file) # loading the data
		except json.decoder.JSONDecodeError:
			balance = dict()
		json_file.seek(0) # going at the beggining of the file			

		if "Stocks" not in balance: # if the file is empty, we add the "Stock" and "total_profit" keys manually
			balance["Stocks"] = {}
			balance["total_profit"] = 0

		for share in profit_dict: 
			# we had the profit and stock traded to the balance for each share in the transaction
			if not share in balance["Stocks"]:
				balance["Stocks"][share] = dict({"profit": 0, "total_stock_traded": 0})

			balance["Stocks"][share]["profit"] += profit_dict[share]["profit"]
			balance["Stocks"][share]["total_stock_traded"] += profit_dict[share]["shares_traded"]

			balance["total_profit"] += profit_dict[share]["profit"]

		json.dump(balance, json_file, indent=4) # writing the dict
		json_file.truncate() # truncating the file

def main(excel_file):

	workbook = openpyxl.load_workbook(excel_file) # loading the workbook
	transactions = workbook.worksheets[1] # getting the transactions

	shares_dict = dict()

	for row in range(4, transactions.max_row):
		security = transactions.cell(column=3, row=row).value
		action   = transactions.cell(column=4, row=row).value
		
		# converting to float for quantity, price and total
		quantity = float(transactions.cell(column=5, row=row).value)
		price    = float(transactions.cell(column=6, row=row).value)
		total    = float(transactions.cell(column=7, row=row).value)

		if security == "Cash": # we ignore Cash deposits
			continue


		_action = "Sold" if action == "Sell" else "Bought"

		if not security in shares_dict:
			shares_dict[security] = {"Bought": [], "Sold": []} # creating the key/value pair in the dict if it doesn't exist
			
		shares_dict[security][_action].append({ # appending the list
				"price": price,
				"shares": quantity
			})
	
	# creating the transaction.json file
	profit_dict = create_transaction_json(shares_dict.copy()) # .copy() so the data is not overriden  
	update_balance_json(profit_dict) # updating the balance.json file

if __name__ == "__main__":

	if not len(sys.argv) == 2: # if the name fo the excel file is not given
		print(f"Usage: python3 {sys.argv[0]} <filename.xlsx>")

	else:
		try:
			main(sys.argv[1]) # lauching the main function
		except Exception as e:
			print(f"The following error occurred: {e}")
		else:
			print("Your JSON files have been created")